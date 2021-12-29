import pandas as pd
import openpyxl
import os
from datetime import datetime

TEMPLATE = 'template/template.xlsx'
COL_TO_CLEAR = []
PARTLIST_ROW = 16
PARTLIST_COL = 2
CAB_NAME_CELL = 'D6'
COL_TO_LIST = {'Amount': 2,
               'Item Code': 3,
               'Item Name': 4,
               'Cost Center': 11}

COL_TO_UNIQUE_LIST = {'Supplier': 35,
                      'Item Code': 38,
                      'Item Name': 39,
                      'Cost Center': 40
                      }
UNIQUE_SHEET = 'unit prices'
UNIQUE_ROW = 5
UNIQUE_COL = 1
SHEET_LIST = [str(i) for i in range(1, 31)]


def clear_df(df: pd.DataFrame, columns: list):
    """
    Posprzątanie df z niepotrzebnych kolumn
    :param df: df do posprzątania
    :param columns: lista kolumn do usunięcia z df
    :return:
    """
    for col in columns:
        df = df.drop([col], axis='columns')

    return df


def save_excel(wb):
    now = datetime.now()
    date = now.strftime("output/%Y%m%d %H%M%S")
    wb.save(filename=f'{date}.xlsx')


class Combine:
    def __init__(self):
        self.files_path = None  # ścieżki do plików .xlsx
        self.all_df = {}  # wszystkie df partlist
        self.unique_df = None  # df unikatów

    def load_to_df(self, fpath):
        """

        :param fpath: plik .xlsx
        :return:
        """
        df = pd.read_excel(fpath)
        cab_name = df.iloc[6, 3]  # nazwa szafy

        for i in range(df.shape[0]):
            if df.iloc[i, 1] == 'Page':
                del df
                df = pd.read_excel(fpath, skiprows=i + 1)
                # df = df[['Amount', 'Item Code', 'Item Name', 'Cost Center']]
                break
        self.all_df[cab_name] = df

    def concat_df(self):
        """
        Połącz all_df w df i usuń duplikaty
        :return:
        """
        self.unique_df = pd.concat([v for v in self.all_df.values()])
        sums = self.unique_df.groupby(['Item Code'], as_index=False)['Amount'].agg('sum')

        self.unique_df = self.unique_df.drop_duplicates(['Item Code'])
        self.unique_df['Amount'] = sums['Amount']

        self.unique_df = self.unique_df.reset_index()
        self.unique_df = self.unique_df.drop(['index'], axis='columns')

    def insert_to_excel(self):
        """
        Zapisz do pliku excel
        :return:
        """
        workbook = openpyxl.load_workbook(TEMPLATE)

        # ----------------------------- #
        # Dotyczy pojedyńczych partlist #
        for sheet_name, df in zip(SHEET_LIST, self.all_df.items()):
            try:
                sheet = workbook[sheet_name]  # wybranie odpowiedniego sheet'a
            except:
                sheet = workbook.create_sheet(sheet_name)
            sheet[CAB_NAME_CELL] = df[0]  # nazwa szafy
            # df_list = df[1].values.tolist()
            # num_row = PARTLIST_ROW  # ustaw początek wpisywania danych (numer wiersza)
            #
            # for row in df_list:
            #     for col, val in enumerate(row, start=PARTLIST_COL):
            #         sheet.cell(row=num_row, column=col).value = val
            #     num_row += 1

            for df_col, col in COL_TO_LIST.items():
                col_list = df[1][df_col].tolist()
                for i, value in enumerate(col_list):
                    sheet.cell(row=i + PARTLIST_ROW, column=col).value = value

            # ----------------------------- #
            # Dotyczy partlisty unikatów -- #
            try:
                sheet = workbook[UNIQUE_SHEET]  # wybranie odpowiedniego sheet'a
            except:
                sheet = workbook.create_sheet(UNIQUE_SHEET)

            # df_list = self.unique_df.values.tolist()
            # num_row = UNIQUE_ROW  # ustaw początek wpisywania danych (numer wiersza)
            # for row in df_list:
            #     for col, val in enumerate(row, start=UNIQUE_COL):
            #         sheet.cell(row=num_row, column=col).value = val
            #     num_row += 1
            for df_col, col in COL_TO_UNIQUE_LIST.items():
                col_list = self.unique_df[df_col].tolist()
                for i, value in enumerate(col_list):
                    sheet.cell(row=i + UNIQUE_ROW, column=col).value = value

        # ----------------------------- #
        # Zapisz do pliku #
        save_excel(workbook)

    def run(self, dir: str):
        """
        :param dir: katalog z partlistami
        :return:
        """
        self.files_path = os.listdir(dir)
        for file in self.files_path:
            if file.lower().endswith('.xlsx'):
                self.load_to_df(os.path.join(dir, file))
        self.concat_df()
        self.insert_to_excel()
